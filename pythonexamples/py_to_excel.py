from openpyxl import load_workbook
import random


class PyToPdf:

    @staticmethod
    def upload_to_excel():
        wb = load_workbook(filename='bs.xlsx')

        ws = wb['Sheet1']

        customer_list = ['Sameer', '16/250 sarabha nagar Extn.', '123456789 RT 0001', '1234 5678 9123',
                         'JMXPS4639K', '9779478123']

        ws.cell(row=12, column=3).value = customer_list[0]
        ws.cell(row=13, column=3).value = customer_list[1]
        ws.cell(row=14, column=3).value = customer_list[2]
        ws.cell(row=15, column=3).value = customer_list[3]
        ws.cell(row=16, column=3).value = customer_list[4]
        ws.cell(row=17, column=3).value = customer_list[5]

        ws.cell(row=12, column=7).value = customer_list[0]
        ws.cell(row=13, column=7).value = customer_list[1]
        ws.cell(row=14, column=7).value = customer_list[2]
        ws.cell(row=15, column=7).value = customer_list[3]
        ws.cell(row=16, column=7).value = customer_list[4]
        ws.cell(row=17, column=7).value = customer_list[5]

        item_list = [['1', '12345', 'mobile phone', '#1001', '2', '10000', '2.5', '2.5', '12'],
                     ['2', '12312', 'something1', '#1002', '4', '100', '2.5', '2.5', '2'],
                     ['3', '12333', 'something2', '#1003', '8', '2000', '2.5', '2.5', '4'],
                     ['4', '12399', 'something3', '#1004', '1', '3000', '2.5', '2.5', '13'],
                     ['5', '12323', 'something4', '#1005', '0', '9000', '2.5', '2.5', '17']]

        ws.cell(row=len(item_list)+24, column=9).value = 'Total'
        ws.cell(row=len(item_list)+25, column=9).value = 'CGST'
        ws.cell(row=len(item_list)+26, column=9).value = 'IGST'
        ws.cell(row=len(item_list)+27, column=9).value = 'Grand Total'

        ws.cell(row=len(item_list) + 30, column=1).value = 'T/C'
        ws.cell(row=len(item_list) + 31, column=1).value = '1. The shipping cost needs to be beared by the seller'
        ws.cell(row=len(item_list) + 32, column=1).value = '2. The seller is not responsible for any damage that happens during the transit'
        ws.cell(row=len(item_list) + 33, column=1).value = '3. If invoice has not been paid in 5 days after due date,'
        ws.cell(row=len(item_list) + 34, column=1).value = '  a tax of 10% of total value is applied to each day of delay'

        ws.cell(row=len(item_list) + 36, column=1).value = 'TRANSPORTATION'
        ws.cell(row=len(item_list) + 37, column=1).value = 'Name:'
        ws.cell(row=len(item_list) + 38, column=1).value = 'GST No:'
        ws.cell(row=len(item_list) + 39, column=1).value = 'Vehicle:'
        ws.cell(row=len(item_list) + 40, column=1).value = 'GST No:'

        ws.cell(row=len(item_list) + 44, column=7).value = 'Proprietor Signature'

        row = 0
        for i in item_list:
            ws.cell(row=22 + row, column=1).value = i[0]
            ws.cell(row=22 + row, column=2).value = i[1]
            ws.cell(row=22 + row, column=3).value = i[2]
            ws.cell(row=22 + row, column=4).value = i[3]
            ws.cell(row=22 + row, column=5).value = i[4]
            ws.cell(row=22 + row, column=6).value = i[5]
            ws.cell(row=22 + row, column=7).value = i[6]
            ws.cell(row=22 + row, column=8).value = i[7]
            ws.cell(row=22 + row, column=10).value = i[8]

            row = row + 1

        name = str(random.randint(1, 100000000))
        wb.save(name + '.xlsx')
        return name + '.xlsx'


# PyToPdf.upload_to_excel()
